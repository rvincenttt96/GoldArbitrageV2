"""Nightly Excel report of paper-trading activity.

Four sheets: a summary, every signal, the signals that could not have been
funded, and why candidates were turned away. The third and fourth exist because
"no trades today" has several very different causes and the summary alone cannot
tell them apart.
"""

from __future__ import annotations

import argparse
import logging
import sqlite3
import sys
from collections import Counter
from datetime import datetime, timedelta, timezone
from datetime import time as dtime
from pathlib import Path

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from services.telegram import BASE_URL, CHAT_ID, TOKEN, TelegramError

log = logging.getLogger("reporting")

PAPER_DB = Path("~/goldarb/paper.db").expanduser()
OUT_DIR = Path("~/goldarb/reports").expanduser()

#: Reports are cut on Iran local time because that is the day the operator
#: lives in, not the server's UTC.
TEHRAN = timezone(timedelta(hours=3, minutes=30))

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _sheet(wb: Workbook, title: str, headers: list[str], first: bool = False):
    ws = wb.active if first else wb.create_sheet()
    ws.title = title
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    return ws


def _autosize(ws) -> None:
    for column in ws.columns:
        width = max((len(str(c.value)) for c in column if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(column[0].column)].width = min(width + 3, 44)


def _local(ts: float) -> str:
    return datetime.fromtimestamp(ts, TEHRAN).strftime("%Y-%m-%d %H:%M:%S")


def day_bounds(day: datetime) -> tuple[float, float]:
    """UTC timestamps spanning one Tehran calendar day."""
    start = datetime.combine(day.date(), dtime.min, tzinfo=TEHRAN)
    return start.timestamp(), (start + timedelta(days=1)).timestamp()


def build_report(conn: sqlite3.Connection, start: float, end: float, out: Path) -> Path:
    signals = conn.execute(
        "SELECT ts, buy_platform, sell_platform, grade, amount_mg, buy_price,"
        " sell_price, buy_fee, sell_fee, net_profit, return_pct, paper_filled,"
        " actionable, block_reason, book_value"
        " FROM signals WHERE ts >= ? AND ts < ? ORDER BY ts",
        (start, end),
    ).fetchall()

    misses = conn.execute(
        "SELECT reason, SUM(n) FROM misses WHERE ts >= ? AND ts < ?"
        " GROUP BY reason ORDER BY SUM(n) DESC",
        (start, end),
    ).fetchall()

    wb = Workbook()

    # -- summary ---------------------------------------------------------
    ws = _sheet(wb, "Summary", ["Metric", "Value"], first=True)
    filled = [s for s in signals if s[11]]
    blocked = [s for s in signals if not s[12]]
    realised = sum(s[9] for s in filled)
    grades = Counter(s[3] for s in signals)

    rows = [
        ("Report date (Tehran)", _local(start)[:10]),
        ("Signals recorded", len(signals)),
        ("Filled on paper", len(filled)),
        ("Blocked by real balances", len(blocked)),
        ("Paper profit (TMN)", realised),
        ("Average net per fill (TMN)", round(realised / len(filled)) if filled else 0),
        ("Best single signal (TMN)", max((s[9] for s in signals), default=0)),
        ("Book value at close (TMN)", signals[-1][14] if signals else 0),
    ]
    for grade in ("A", "B", "C", "D"):
        rows.append((f"Grade {grade} signals", grades.get(grade, 0)))
    for label, value in rows:
        ws.append([label, value])
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 22

    # -- every signal ----------------------------------------------------
    ws = _sheet(wb, "Signals", [
        "Time (Tehran)", "Buy", "Sell", "Grade", "Size (g)", "Buy price",
        "Sell price", "Buy fee", "Sell fee", "Net (TMN)", "Return %",
        "Paper filled", "Actionable", "Blocked because", "Book value",
    ])
    for s in signals:
        ws.append([
            _local(s[0]), s[1], s[2], s[3], s[4] / 1000, s[5], s[6], s[7], s[8],
            s[9], round(s[10], 3), "yes" if s[11] else "no",
            "yes" if s[12] else "no", s[13], s[14],
        ])
    _autosize(ws)

    # -- signals we could not fund ---------------------------------------
    ws = _sheet(wb, "Not actionable", [
        "Time (Tehran)", "Route", "Grade", "Size (g)", "Net foregone (TMN)", "Reason",
    ])
    for s in blocked:
        ws.append([_local(s[0]), f"{s[1]} -> {s[2]}", s[3], s[4] / 1000, s[9], s[13]])
    if blocked:
        ws.append([])
        ws.append(["", "TOTAL FOREGONE", "", "", sum(s[9] for s in blocked), ""])
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
    _autosize(ws)

    # -- why candidates were turned away ---------------------------------
    ws = _sheet(wb, "Rejections", ["Reason", "Count", "Share %"])
    total = sum(n for _, n in misses) or 1
    for reason, n in misses:
        ws.append([reason, n, round(n / total * 100, 2)])
    _autosize(ws)

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


def send_to_telegram(path: Path, caption: str) -> None:
    if not TOKEN or not CHAT_ID:
        raise TelegramError("TELEGRAM_TOKEN and TELEGRAM_CHAT_ID are not set")

    with path.open("rb") as handle:
        response = requests.post(
            f"{BASE_URL}/bot{TOKEN}/sendDocument",
            data={"chat_id": CHAT_ID, "caption": caption, "parse_mode": "HTML"},
            files={"document": (path.name, handle)},
            timeout=120,
        )
    response.raise_for_status()
    payload = response.json()
    if payload.get("ok") is False:
        raise TelegramError(f"Telegram rejected the document: {payload}")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--paper-db", type=Path, default=PAPER_DB)
    parser.add_argument("--out-dir", type=Path, default=OUT_DIR)
    parser.add_argument("--date", help="YYYY-MM-DD in Tehran time; defaults to today")
    parser.add_argument("--send", action="store_true")
    args = parser.parse_args()

    logging.basicConfig(
        level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s",
        stream=sys.stdout,
    )

    day = (
        datetime.strptime(args.date, "%Y-%m-%d").replace(tzinfo=TEHRAN)
        if args.date
        else datetime.now(TEHRAN)
    )
    start, end = day_bounds(day)

    conn = sqlite3.connect(f"file:{args.paper_db}?mode=ro", uri=True, timeout=30)
    out = args.out_dir / f"goldarb-{day.strftime('%Y-%m-%d')}.xlsx"
    build_report(conn, start, end, out)
    conn.close()

    log.info("wrote %s (%.1f KB)", out, out.stat().st_size / 1024)

    if args.send:
        send_to_telegram(out, f"<b>Daily paper trading report</b>\n{day:%Y-%m-%d}")
        log.info("sent to telegram")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
